#!/usr/bin/env python3
"""Build the standard MRA Project Intake Template (.xlsx).

Slim header (logo + title) + a tiny info block (Project / Job # / PM), then a task
table that MIRRORS the master 'Project Tasks' sheet — same columns AND the same look
(dark header, gold milestone rows) so an upload/paste is an easy match.

Task table = master cols A..O:
    A=Project  B=Phase  C=Type  D=Task  E=Start  F=Finish  G=Duration
    H=Assigned To  I=Status  J=PM  K=Milestone  L=Comments  M=Task ID  N=Predecessor
    O=Sub

Only Project / Client and Project Manager are functionally required (Project & PM
auto-fill down columns A & J; the importer also falls back to these). MRA Job # is
optional (appended to the project name). Task ID (M) is system-assigned on import;
Predecessor (N) is by Task ID — both shown greyed. Sub (O) marks a row as a SUBTASK
of the task above it: put an "x" there and the dashboard nests it under its parent
(MS-Project style, collapsible). Ships with 10 gold milestone rows (Milestone 1..10)
each followed by 3 blank task rows + 2 indented blank subtask rows (Sub pre-set to
"x") so the parent → subtask pattern is built in and ready to fill.

Import-Intake.ps1 reads this back (MIRROR layout: header A="Project" & D="Task").
Change a label/column here -> update the matching reader in Import-Intake.ps1.
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.drawing.image import Image as XLImage

HERE = os.path.dirname(os.path.abspath(__file__))
ORANGE = "E24E26"
NAVY = "1F3864"    # dark header band — matches the master Project Tasks sheet
GOLD = "FFE699"    # milestone row shading — matches the master
DARK = "1F2430"
GREY = "6B7280"
FIELD = "FFF7F4"   # very light orange tint for fill-in cells
AUTO = "F3F4F6"    # light grey for auto-filled / system columns
SUB_FILL = "EEF2F7"   # very light blue-grey — subtask slots + the subtask note banner

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Enter Here"

# ---- Lists sheet (dropdown sources; Project is intentionally NOT here) ------
lst = wb.create_sheet("Lists")
LISTS = {
    "B": ("Phase", ["Project Planning", "Creative Design", "Design & Detail",
                    "Fabrication", "Production",
                    "Electrical & Technology", "Graphics",
                    "Post Production", "Launch"]),
    "C": ("Type", ["Meeting", "Hard Date", "Design Task", "Materials",
                   "Production Task", "Client", "Logistics", "QA",
                   "Milestone"]),
    "D": ("Status", ["Not Started", "In Progress", "Completed", "On Hold"]),
    "E": ("Milestone", ["Yes", "No"]),
    "F": ("PM", ["Megan Fraser", "Al Karloff", "Stephanie Hardie", "Alex Karam",
                 "Luciana Giglio", "Frank Mancina", "Mark St Jean",
                 "Sherri Washington", "Sherrill Buchan", "Cindy Irland",
                 "Mitch Schirr", "Heather Maloney", "TBD"]),
    "G": ("Assigned", ["MRA", "MRA Shop", "Combined Effort", "Megan Fraser",
                       "Al Karloff", "Gino Bitonti", "Sal", "Doug",
                       "MasterWraps", "Electricians", "Vendor", "Medtronic",
                       "Learning Undefeated", "IXL/TSS", "Brinkbit",
                       "Siemens DI", "Heitek", "Other"]),
    "H": ("Sub", ["x"]),
}
for col, (head, vals) in LISTS.items():
    lst[f"{col}1"] = head
    for i, v in enumerate(vals):
        lst[f"{col}{i+2}"] = v
lst.sheet_state = "veryHidden"

# ---- Column widths (task table mirrors master A..O) -------------------------
widths = {"A": 26, "B": 18, "C": 13, "D": 42, "E": 11, "F": 11, "G": 9,
          "H": 20, "I": 13, "J": 15, "K": 10, "L": 26, "M": 9, "N": 12, "O": 7}
for c, w in widths.items():
    ws.column_dimensions[c].width = w

thin = Side(style="thin", color="D1D5DB")
box = Border(left=thin, right=thin, top=thin, bottom=thin)
bottom = Border(bottom=Side(style="thin", color="C9CDD3"))

# ---- Slim letterhead (logo + title only) -----------------------------------
logo_path = os.path.join(HERE, "logo.png")
if os.path.exists(logo_path):
    img = XLImage(logo_path)
    img.width = 74
    img.height = 74
    ws.add_image(img, "A1")

for r, h in {1: 22, 2: 20, 3: 6}.items():
    ws.row_dimensions[r].height = h

ws.merge_cells("C1:O1")
t = ws["C1"]
t.value = "PROJECT INTAKE  —  TASK SCHEDULE"
t.font = Font(bold=True, size=18, color=ORANGE)
t.alignment = Alignment(vertical="center")
ws.merge_cells("C2:O2")
a = ws["C2"]
a.value = "950 E Whitcomb Ave  ·  Madison Heights, MI 48071  ·  248.629.2929"
a.font = Font(size=9, color=GREY)

# ---- Tiny info block: Project / Job # / PM only ----------------------------
def label(ref, text):
    c = ws[ref]
    c.value = text
    c.font = Font(bold=True, size=10, color=DARK)
    c.alignment = Alignment(horizontal="right", vertical="center")

def field(top_left, span_to, prompt=None, title=None):
    ws.merge_cells(f"{top_left}:{span_to}")
    c = ws[top_left]
    c.fill = PatternFill("solid", fgColor=FIELD)
    c.border = bottom
    c.alignment = Alignment(vertical="center", indent=1)
    c.font = Font(size=11, color=DARK)
    if prompt:
        dv = DataValidation(allow_blank=True, showInputMessage=True,
                            showErrorMessage=False, prompt=prompt, promptTitle=title or "")
        ws.add_data_validation(dv)
        dv.add(top_left)
    return c

INFO0 = 4   # Project field = B4, PM field = B5 (referenced by the autofill formulas)
for rr in (INFO0, INFO0 + 1):
    ws.row_dimensions[rr].height = 20

label(f"A{INFO0}", "Project / Client:")
field(f"B{INFO0}", f"D{INFO0}", title="Project / Client",
      prompt="Type the project name (new jobs welcome). It auto-fills down column A.")
label(f"F{INFO0}", "MRA Job #:")
field(f"G{INFO0}", f"H{INFO0}", title="MRA Job #",
      prompt='Job number if you have one, or type "NEW".')

label(f"A{INFO0+1}", "Project Manager:")
field(f"B{INFO0+1}", f"D{INFO0+1}", title="Project Manager",
      prompt="The PM for this project — auto-fills down column J.")

# ---- "How to" note banners (so people SEE & understand the two special rows) -
def note_banner(r, text, fill):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=15)
    c = ws.cell(row=r, column=1, value=text)
    c.fill = PatternFill("solid", fgColor=fill)
    c.font = Font(size=10, bold=True, color=DARK)
    c.alignment = Alignment(vertical="center", wrap_text=True, indent=1)
    for cc in range(1, 16):
        ws.cell(row=r, column=cc).border = box
    ws.row_dimensions[r].height = 26

note_banner(INFO0 + 2,
    "◆  MILESTONES (the gold rows):  set  Milestone = Yes  for a key date / hard deadline "
    "— they print on the schedule and show as ◆ gates on the dashboard.",
    GOLD)
note_banner(INFO0 + 3,
    "↳  SUBTASKS:  put an  x  in the  Sub  column (far right →)  to make a row a sub-step "
    "of the task ABOVE it — it shows indented & collapsible on the board. A few blank x-marked "
    "subtask slots are already built in under each section.",
    SUB_FILL)

# ---- Task table header (dark band, like the master sheet) -------------------
HDR = INFO0 + 5   # rows: 4-5 info · 6 milestone note · 7 subtask note · 8 spacer · 9 header
ws.row_dimensions[HDR - 1].height = 8
heads = ["Project", "Phase", "Type", "Task", "Start", "Finish", "Duration",
         "Assigned To", "Status", "PM", "Milestone", "Comments", "Task ID",
         "Predecessor", "Sub"]
LEFT_COLS = {1, 2, 4, 8, 10, 12}
for c, h in enumerate(heads, start=1):
    cell = ws.cell(row=HDR, column=c, value=h)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(vertical="center",
                               horizontal="left" if c in LEFT_COLS else "center",
                               indent=1 if c in LEFT_COLS else 0)
    cell.border = box
ws.row_dimensions[HDR].height = 24

# ---- Body: 10 gold milestone rows, each followed by 3 blank task rows + 2 ----
#      indented blank SUBTASK rows (Sub pre-set to "x") so the parent → subtask
#      pattern is built in. Within each block (offsets after the milestone):
#         1,2,3 = normal task slots   ·   4,5 = subtask slots (indented, Sub=x)
DATA0 = HDR + 1
NMILE, GAP = 10, 5
STEP = GAP + 1
mile_at = {DATA0 + i * STEP: i + 1 for i in range(NMILE)}
sub_rows = set()
for i in range(NMILE):
    base = DATA0 + i * STEP
    sub_rows.add(base + 4); sub_rows.add(base + 5)     # last 2 of each block
LASTROW = DATA0 + NMILE * STEP + 60
auto_fill = PatternFill("solid", fgColor=AUTO)
gold_fill = PatternFill("solid", fgColor=GOLD)
sub_fill = PatternFill("solid", fgColor=SUB_FILL)
projF = '=IF($B$%d="","",$B$%d)' % (INFO0, INFO0)
pmF = '=IF($B$%d="","",$B$%d)' % (INFO0 + 1, INFO0 + 1)
sub_prompt = ('Subtask of the task above. Type the subtask here and keep the '
              '"x" in the Sub column — the dashboard nests it (collapsible) '
              'under its parent. Clear the "x" to make it a normal task.')
sub_dv = DataValidation(allow_blank=True, showInputMessage=True,
                        showErrorMessage=False, prompt=sub_prompt,
                        promptTitle="Subtask")
ws.add_data_validation(sub_dv)

for r in range(DATA0, LASTROW):
    for c in range(1, 16):     # A..O
        ws.cell(row=r, column=c).border = box
    ws.cell(row=r, column=1, value=projF)      # A Project (autofill)
    ws.cell(row=r, column=10, value=pmF)       # J PM (autofill)
    ws.cell(row=r, column=5).number_format = "m/d/yyyy"   # Start
    ws.cell(row=r, column=6).number_format = "m/d/yyyy"   # Finish
    if r in mile_at:
        n = mile_at[r]
        ws.cell(row=r, column=3, value="Milestone")        # C Type
        ws.cell(row=r, column=4, value="Milestone %d" % n) # D Task
        ws.cell(row=r, column=9, value="Not Started")      # I Status
        ws.cell(row=r, column=11, value="Yes")             # K Milestone
        for c in range(1, 16):
            ws.cell(row=r, column=c).fill = gold_fill
    else:
        ws.cell(row=r, column=1).fill = auto_fill          # A grey (autofill)
        ws.cell(row=r, column=10).fill = auto_fill         # J grey (autofill)
        ws.cell(row=r, column=13).fill = auto_fill         # M Task ID (system)
        ws.cell(row=r, column=14).fill = auto_fill         # N Predecessor (system)
        if r in sub_rows:
            ws.cell(row=r, column=4).alignment = Alignment(indent=3)   # D indent
            ws.cell(row=r, column=15, value="x")           # O Sub = x (pre-set)
            sub_dv.add(ws.cell(row=r, column=4))           # click-prompt on Task
            for c in (4, 15):
                ws.cell(row=r, column=c).fill = sub_fill

# ---- Dropdowns -------------------------------------------------------------
def add_dv(col_letter, src, last=LASTROW - 1):
    dv = DataValidation(type="list", formula1=src, allow_blank=True,
                        showErrorMessage=False, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{DATA0}:{col_letter}{last}")

add_dv("B", "=Lists!$B$2:$B$10")   # Phase
add_dv("C", "=Lists!$C$2:$C$10")   # Type
add_dv("H", "=Lists!$G$2:$G$19")   # Assigned To
add_dv("I", "=Lists!$D$2:$D$5")    # Status
add_dv("K", "=Lists!$E$2:$E$3")    # Milestone
add_dv("O", "=Lists!$H$2:$H$2")    # Sub (x / blank)
dv_pm = DataValidation(type="list", formula1="=Lists!$F$2:$F$14",
                       allow_blank=True, showErrorMessage=False, showDropDown=False)
ws.add_data_validation(dv_pm)
dv_pm.add(f"B{INFO0+1}")

# ---- Final touches ----------------------------------------------------------
ws.freeze_panes = f"A{DATA0}"
ws.sheet_view.showGridLines = False
ws.print_options.horizontalCentered = True
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_margins.left = ws.page_margins.right = 0.4
ws.page_margins.top = ws.page_margins.bottom = 0.5

out = os.path.join(HERE, "MRA_Project_Intake_Template.xlsx")
wb.save(out)
print("wrote", out, "| header row", HDR, "| data starts", DATA0,
      "| milestones at rows", sorted(mile_at.keys()))
